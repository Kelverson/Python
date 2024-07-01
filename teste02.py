import pyodbc
#import xlsxwriter
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

# Parâmetros de conexão
server = ' '
database = ' '
username = ' '
password = ' '

# String de conexão
conn_str = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'

# Conectar ao banco de dados
conn = pyodbc.connect(conn_str)

# Criar um cursor
cursor = conn.cursor()

# Abrir o arquivo de texto em modo de leitura
with open('C:/temp/BUSCA.txt', 'r') as file:
    # Ler linhas do arquivo e remover espaços em branco
    b1_cod = [line.strip() for line in file.readlines()]

print(b1_cod)

TamanhoArray = len(b1_cod)
contador = 0
p = 0

def NovaConsulta(resultado):
    j = 0
    f = 0
    teste = 0
    tamanhoArray = len(resultado)
    #print(resultado)
    #print("1")
    resultadosFunction = []
    #print(resultadosFunction)
    resultadosFunction = resultado.copy()
    continuar = True
    while continuar:
        for i in range(len(resultado)):
            # Executar a consulta
            query = """
                SELECT
                    CASE
                        WHEN SB1.B1_XCATNL = 1 THEN 'EM LINHA'
                        WHEN SB1.B1_XCATNL = 2 THEN 'FORA DE LINHA'
                        ELSE 'SEM MOVIMENTO'
                    END AS 'CATEG. NL',
                    RTRIM(SB1.B1_COD) AS 'CODIGO',
                    RTRIM(SB1.B1_DESC) AS 'DESC_PRODUTO',
                    RTRIM(SB1.B1_TIPO) AS 'TIPO',
                    RTRIM(Z4_FAMILIA) AS 'COD_FAMILIA',
                    RTRIM(SZ4.Z4_DESCFAM) AS 'DESC_FAMILIA_NL',
                    RTRIM(SB12.B1_COD) AS 'COD_COMP',
                    RTRIM(SB12.B1_DESC) AS 'DESC_COMP',
                    CASE
                        WHEN SB1.B1_VIGENC <> '' THEN FORMAT(CONVERT(DATETIME2,SB1.B1_VIGENC),'dd/MM/yyyy')
                        ELSE ''
                    END AS 'DATA_VIGENC'
                FROM SG1010 AS SG1
                INNER JOIN SB1010 AS SB1 ON SB1.B1_COD = SG1.G1_COD
                    AND SB1.B1_FILIAL = SG1.G1_FILIAL
                    AND G1_REVFIM IN (SB1.B1_REVATU,'ZZZ','YYY')
                    AND SB1.D_E_L_E_T_ = ''
                INNER JOIN SB1010 AS SB12 ON G1_COMP = SB12.B1_COD
                    AND G1_FILIAL = SB12.B1_FILIAL
                    AND SB12.D_E_L_E_T_ = ''
                LEFT JOIN SZ4010 AS SZ4 ON SZ4.D_E_L_E_T_ = ''
                    AND Z4_FILIAL = SB1.B1_FILIAL
                    AND Z4_FAMILIA IN (' ', SB1.B1_XFAMILI)
                    AND Z4_GRUPO IN ('', SB1.B1_GRUPO)
                WHERE SG1.D_E_L_E_T_ = ''
                    AND G1_FILIAL = '0201'
                    AND SB1.B1_XFAMILI IN (Z4_FAMILIA,'') 
                    AND SB1.B1_XPORTIF IN (Z4_PORTIFO,'')
                    AND SB1.B1_XSBPORT IN (Z4_SBPORTI,'')
                    AND SB12.B1_COD = ?
                ORDER BY SB1.B1_VIGENC DESC
            """
            if(j < tamanhoArray):
                B1_COD = resultado[j][1]
                j+=1
            elif(f < teste ):
                #print(f)
                B1_COD = resultadosFunction[f][1]
                #print(B1_COD)
                f+=1
            # Executar a consulta com o parâmetro
            cursor.execute(query, (B1_COD,))

            for row in cursor.fetchall():
                resultadosFunction.append(row)
            k = len(resultadosFunction) - 1

            teste = len(resultadosFunction)

            if (resultadosFunction[-1][3] in ['PA','BN']):
                continuar = False

    # Crie um objeto Workbook
    wb = openpyxl.Workbook()
    # Crie uma nova planilha
    ws = wb.active
    # Conteúdo antes dos headers
    conteudo_extra = ['PRODUTO PESQUISADO: ' + sb12_b1_cod]
    # Adicione o conteúdo extra na primeira linha
    for i, valor in enumerate(conteudo_extra, start=1):
        ws.cell(row=i, column=1).value = valor
    # Headers
    headers = ['CATEG. NL', 'CODIGO', 'DESC_PRODUTO', 'TIPO', 'COD_FAMILIA', 'DESC_FAMILIA_NL', 'COD_COMP', 'DESC_COMP', 'DATA_VIGENC']
    # Adicione os headers
    for col, header in enumerate(headers, start=1):
        ws.cell(row=len(conteudo_extra) + 1, column=col).value = header
    # Escrever os dados
    for row_idx, row in enumerate(resultadosFunction, start=len(conteudo_extra) + 2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value
    # Definir a formatação como tabela a partir da segunda linha
    tab = Table(displayName="Tabela1", ref="A2:I{}".format(len(resultadosFunction) + len(conteudo_extra) + 1))
    # Estilo da tabela
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(caminho_arquivo)
    print("Resultados exportados para Excel com sucesso!")

while p < TamanhoArray:
    
    #Vetor Para salvar os resultados da primeira consulta. 
    resultados = []
    
    # Parâmetro para SB12.B1_COD
    sb12_b1_cod = b1_cod[p]
    p+=1
    # Especificar o caminho onde salvar o arquivo
    caminho_arquivo = 'C:/temp/resultados'+sb12_b1_cod+'.xlsx'
    # Executar a consulta
    query = """
        SELECT
            CASE
                WHEN SB1.B1_XCATNL = 1 THEN 'EM LINHA'
                WHEN SB1.B1_XCATNL = 2 THEN 'FORA DE LINHA'
                ELSE 'SEM MOVIMENTO'
            END AS 'CATEG. NL',
            RTRIM(SB1.B1_COD) AS 'CODIGO',
            RTRIM(SB1.B1_DESC) AS 'DESC_PRODUTO',
            RTRIM(SB1.B1_TIPO) AS 'TIPO',
            RTRIM(Z4_FAMILIA) AS 'COD_FAMILIA',
            RTRIM(SZ4.Z4_DESCFAM) AS 'DESC_FAMILIA_NL',
            RTRIM(SB12.B1_COD) AS 'COD_COMP',
            RTRIM(SB12.B1_DESC) AS 'DESC_COMP',
            CASE
                WHEN SB1.B1_VIGENC <> '' THEN FORMAT(CONVERT(DATETIME2,SB1.B1_VIGENC),'dd/MM/yyyy')
                ELSE ''
            END AS 'DATA_VIGENC'
        FROM SG1010 AS SG1
        INNER JOIN SB1010 AS SB1 ON SB1.B1_COD = SG1.G1_COD
            AND SB1.B1_FILIAL = SG1.G1_FILIAL
            AND G1_REVFIM IN (SB1.B1_REVATU,'ZZZ','YYY')
            AND SB1.D_E_L_E_T_ = ''
        INNER JOIN SB1010 AS SB12 ON G1_COMP = SB12.B1_COD
            AND G1_FILIAL = SB12.B1_FILIAL
            AND SB12.D_E_L_E_T_ = ''
        LEFT JOIN SZ4010 AS SZ4 ON SZ4.D_E_L_E_T_ = ''
            AND Z4_FILIAL = SB1.B1_FILIAL
            AND Z4_FAMILIA IN (' ', SB1.B1_XFAMILI)
            AND Z4_GRUPO IN ('', SB1.B1_GRUPO)
        WHERE SG1.D_E_L_E_T_ = ''
            AND G1_FILIAL = '0201'
            AND SB1.B1_XFAMILI IN (Z4_FAMILIA,'') 
            AND SB1.B1_XPORTIF IN (Z4_PORTIFO,'')
            AND SB1.B1_XSBPORT IN (Z4_SBPORTI,'')
            AND SB12.B1_COD = ?
        ORDER BY SB1.B1_VIGENC DESC
    """
    
    # Executar a consulta com o parâmetro
    cursor.execute(query, (sb12_b1_cod,))
    #print(cursor.fetchall())
    #print(cursor.rowcount)
    if(cursor.rowcount < 0 or cursor.rowcount > 0):
        for row in cursor.fetchall():
            resultados.append(row)

        #print(resultados[0]) 
        NovaConsulta(resultados)
    else:
        # Crie um objeto Workbook
        wb = openpyxl.Workbook()
        # Crie uma nova planilha
        ws = wb.active
        # Conteúdo antes dos headers
        conteudo_extra = ['PRODUTO PESQUISADO: ' + sb12_b1_cod]
        # Adicione o conteúdo extra na primeira linha
        for i, valor in enumerate(conteudo_extra, start=1):
            ws.cell(row=i, column=1).value = valor
        # Headers
        headers = ['PRODUTO NÃO CONSTA EM ESTRUTURAS']
        # Adicione os headers
        for col, header in enumerate(headers, start=1):
            ws.cell(row=len(conteudo_extra) + 1, column=col).value = header
            
        wb.save(caminho_arquivo)
        print("Resultados exportados para Excel com sucesso!")

conn.close()
