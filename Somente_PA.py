import pyodbc
import os
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter as tk
from tkinter import filedialog

root = tk.Tk()
root.withdraw()  # Esconde a janela principal

# Abre a caixa de diálogo para selecionar um arquivo
arquivo_busca = filedialog.askopenfilename(title="Abrir arquivo com os códigos!!", filetypes=[("Text files", "*.txt")])

# Parâmetros de conexão
server = '192.168.0.8'
database = 'TOTVS_NEWLINE_PRD'
username = 'PCP'
password = 'Gnl@2030'

# String de conexão
conn_str = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'

# Conectar ao banco de dados
conn = pyodbc.connect(conn_str)

# Criar um cursor
cursor = conn.cursor()

# Abrir o arquivo de texto em modo de leitura
with open(arquivo_busca, 'r') as file:
    # Ler linhas do arquivo, remover espaços em branco e ignorar linhas vazias
    b1_cod = [line.strip() for line in file.readlines() if line.strip()]

print(b1_cod)

TamanhoArray = len(b1_cod)
contador = 0
p = 0

root.withdraw()  # Esconde a janela principal
file_path = filedialog.askdirectory(title="Onde deseja salvar!!")
file_path2 = file_path

def NovaConsulta(resultado):
    j = 0
    f = 0
    teste = 0
    tamanhoArray = len(resultado)

    resultadosFunction = []
    # Inicializar um conjunto para armazenar códigos únicos
    codigos_unicos = set()
    resultadosFunction = resultado.copy()
    continuar = True

    while continuar:
        for i in range(len(resultado)):
            # Executar a consulta
            query = """
                SELECT
                    CASE
                        WHEN SB1.B1_XCATNL = 1 THEN 'EM LINHA'
                        WHEN SB1.B1_XCATNL = 2 THEN 'FORA DE LINHA'
		                WHEN SB1.B1_XCATNL = 3 THEN 'SEM MOVIMENTO'
		                WHEN SB1.B1_XCATNL = 4 THEN 'ANALISE FL'
                        ELSE 'SUBSTIUIÇÃO'
                    END AS 'CATEG. NL',
                    RTRIM(SB1.B1_COD) AS 'CODIGO',
                    RTRIM(SB1.B1_DESC) AS 'DESC_PRODUTO',
                    RTRIM(SB1.B1_TIPO) AS 'TIPO',
                    RTRIM(Z4_FAMILIA) AS 'COD_FAMILIA',
                    RTRIM(SZ4.Z4_DESCFAM) AS 'DESC_FAMILIA_NL',
                    RTRIM(SB12.B1_COD) AS 'COD_COMP',
                    RTRIM(SB12.B1_DESC) AS 'DESC_COMP',
                    RTRIM(SG1.G1_QUANT) AS 'QUANT_COMP',
                    CASE
                        WHEN SB1.B1_VIGENC <> '' THEN FORMAT(CONVERT(DATETIME2,SB1.B1_VIGENC),'dd/MM/yyyy')
                        ELSE ''
                    END AS 'DATA_VIGENC'
                FROM SG1010 AS SG1 WITH (NOLOCK)

                INNER JOIN SB1010 AS SB1 WITH (NOLOCK) ON SB1.B1_COD = SG1.G1_COD
                    AND SB1.B1_FILIAL = SG1.G1_FILIAL
                    AND G1_REVFIM IN (SB1.B1_REVATU,'ZZZ','YYY')
                    AND SB1.D_E_L_E_T_ = ''
                INNER JOIN SB1010 AS SB12 WITH (NOLOCK) ON G1_COMP = SB12.B1_COD
                    AND G1_FILIAL = SB12.B1_FILIAL
                    AND SB12.D_E_L_E_T_ = ''
                LEFT JOIN SZ4010 AS SZ4 WITH (NOLOCK) ON SZ4.D_E_L_E_T_ = ''
                    AND Z4_FILIAL = SB1.B1_FILIAL
                    AND Z4_FAMILIA IN (' ', SB1.B1_XFAMILI)
                    AND Z4_GRUPO IN ('', SB1.B1_GRUPO)
                WHERE SG1.D_E_L_E_T_ = ''
                    AND G1_FILIAL = '0201'
                    AND SB1.B1_XFAMILI IN (Z4_FAMILIA,'') 
                    AND SB1.B1_XPORTIF IN (Z4_PORTIFO,'')
                    AND SB1.B1_XSBPORT IN (Z4_SBPORTI,'')
                    AND SB12.B1_COD = ?
                ORDER BY SB1.B1_VIGENC DESC
            """
            #AC000405
            if(j < tamanhoArray):
                B1_COD = resultado[j][1]
                j+=1
            elif(f < teste ):
                
                B1_COD = resultadosFunction[f][1]
                
                f+=1
            # Executar a consulta com o parâmetro
            cursor.execute(query, (B1_COD,))

            for row in cursor.fetchall():
                if row[1] not in codigos_unicos:
                    resultadosFunction.append(row)
                    codigos_unicos.add(row[1])

            teste = len(resultadosFunction)
            #print(teste)

            if (f == teste):
            #if (resultadosFunction[-1][3] != ''):
            #if (resultadosFunction[-1][3] in ['PA','BN','PP']):
                continuar = False
    a = 0
    # Estrutura de repetição para trocar os dados componente vindo da query pelos dados do insumo pesquisado
    # Onde trocamos os valores dos campos COD_COMP, DESC_COMP, QUAND_COMP e DATA_VIGENC, para os valores do insumo
    for cod in resultado:
        for a in range(len(resultadosFunction)):
            
            if(resultadosFunction[a][6] == cod[1]):
                #Condição onde verifica se a quantidade do material PP é menor que o Insumo, 
                #Se for TRUE, a quantidade do material PP é TRocada pela quantidade do Insumo
                #Se for FALSE, a quantidade do material PP é mostrada no relatório
                #print(cod[1])
                #print('posicao A : ',resultadosFunction[a])

                b = 0
                for b in range(len(resultado)):
                    if(resultado[b][1] == cod[1]):
                        #print(b)
                        #print(resultado[b])
                        num = float(resultado[b][8])
                        if( num < 1 ):
                            mult = float(resultadosFunction[a][8]) * float(resultado[b][8])
                            resultadosFunction[a][8] = str(mult)
                            #print(mult)
                        elif(resultadosFunction[a][8] < resultado[b][8] and resultadosFunction[a][8] == '1'):

                            resultadosFunction[a][8] = resultado[b][8]

                        elif(resultadosFunction[a][8] != '1' and resultado[b][8] != '1'):

                            mult = float(resultadosFunction[a][8]) * float(resultado[b][8])
                            resultadosFunction[a][8] = str(mult)
                            #print(mult)

                        resultadosFunction[a][6] = resultado[b][6]
                        resultadosFunction[a][7] = resultado[b][7]
                        #resultadosFunction[a][8] = resultado[f][8]
                        resultadosFunction[a][9] = resultado[b][9]
                        #print(resultadosFunction[a][8])

    c = 0
    
    for cod in resultadosFunction:

        for c in range(len(resultadosFunction)):
            
            if(resultadosFunction[c][6] == cod[1]):
                #Condição onde verifica se a quantidade do material PP é menor que o Insumo, 
                #Se for TRUE, a quantidade do material PP é TRocada pela quantidade do Insumo
                #Se for FALSE, a quantidade do material PP é mostrada no relatório 
                #print(cod[1])
                #print('Posicao C',resultadosFunction[c])
                e = 0
                for e in range(len(resultadosFunction)):
                    if(resultadosFunction[e][1] == cod[1]):
                        #print('Posicao E',resultadosFunction[e])
                        #os.system('cls')
                        num = float(resultadosFunction[e][8])
                        if( num < 1 and resultadosFunction[c][8] == '1'):
                            num = float(resultadosFunction[e][8]) * float(resultadosFunction[c][8])
                            resultadosFunction[c][8] = str(num)

                        elif(resultadosFunction[c][8] != '1' and resultadosFunction[e][8] != '1'):
                            num = float(resultadosFunction[e][8]) * float(resultadosFunction[c][8])
                            resultadosFunction[c][8] = str(num)

                        resultadosFunction[c][6] = resultadosFunction[e][6]
                        resultadosFunction[c][7] = resultadosFunction[e][7]
                        #resultadosFunction[a][8] = resultado[f][8]
                        resultadosFunction[c][9] = resultadosFunction[e][9]
    # Crie um objeto Workbook
    wb = openpyxl.Workbook()
    # Crie uma nova planilha
    ws = wb.active
    # Conteúdo antes dos headers
    conteudo_extra = ['PRODUTO PESQUISADO: ' + sb12_b1_cod]
    # Adicione o conteúdo extra na primeira linha
    for i, valor in enumerate(conteudo_extra, start=1):
        ws.cell(row=i, column=1).value = valor
    # Headers
    headers = ['CATEG. NL', 'CODIGO', 'DESC_PRODUTO', 'TIPO', 'COD_FAMILIA', 'DESC_FAMILIA_NL', 'COD_COMP', 'DESC_COMP', 'QUANT_COMP', 'DATA_VIGENC']
    # Adicione os headers
    for col, header in enumerate(headers, start=1):
        ws.cell(row=len(conteudo_extra) + 1, column=col).value = header
        
    # Contador para manter o controle da linha atual na planilha
    current_row_idx = len(conteudo_extra) + 2
    
    # Escrever os dados
    for row in resultadosFunction:
        if row[3] == "PA" or row[1].startswith('NAC'):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=current_row_idx, column=col_idx).value = value
            # Incrementar o índice da linha apenas quando uma linha é escrita
            current_row_idx += 1

    # Definir a formatação como tabela a partir da segunda linha
    tab = Table(displayName="Tabela1", ref="A2:J{}".format(current_row_idx - 1))
    # Estilo da tabela
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab) 
    #print(file_path)
    # Especificar o caminho onde salvar o arquivo
    wb.save(file_path)
    print("Resultados "+nomeArquivo+" exportados para Excel com sucesso!")

while p < TamanhoArray:
    
    #Vetor Para salvar os resultados da primeira consulta. 
    resultados = []
    
    # Parâmetro para SB12.B1_COD
    sb12_b1_cod = b1_cod[p]
    p+=1
    nomeArquivo = sb12_b1_cod.replace('/','_')
    #caminho_arquivo = 'C:/temp/resultados'+nomeArquivo+'.xlsx'
    file_path = file_path2+"/"+nomeArquivo+'.xlsx'
    # Executar a consulta
    query = """
        SELECT
                CASE
                    WHEN SB1.B1_XCATNL = 1 THEN 'EM LINHA'
                    WHEN SB1.B1_XCATNL = 2 THEN 'FORA DE LINHA'
		            WHEN SB1.B1_XCATNL = 3 THEN 'SEM MOVIMENTO'
		            WHEN SB1.B1_XCATNL = 4 THEN 'ANALISE FL'
                    ELSE 'SUBSTIUIÇÃO'
                END AS 'CATEG. NL',
                RTRIM(SB1.B1_COD) AS 'CODIGO',
                RTRIM(SB1.B1_DESC) AS 'DESC_PRODUTO',
                RTRIM(SB1.B1_TIPO) AS 'TIPO',
                RTRIM(Z4_FAMILIA) AS 'COD_FAMILIA',
                RTRIM(SZ4.Z4_DESCFAM) AS 'DESC_FAMILIA_NL',
                RTRIM(SB12.B1_COD) AS 'COD_COMP',
                RTRIM(SB12.B1_DESC) AS 'DESC_COMP',
                RTRIM(SG1.G1_QUANT) AS 'QUANT_COMP',
                CASE
                    WHEN SB1.B1_VIGENC <> '' THEN FORMAT(CONVERT(DATETIME2,SB1.B1_VIGENC),'dd/MM/yyyy')
                    ELSE ''
                END AS 'DATA_VIGENC'
            FROM SG1010 AS SG1 WITH (NOLOCK)
            INNER JOIN SB1010 AS SB1 WITH (NOLOCK) ON SB1.B1_COD = SG1.G1_COD
                AND SB1.B1_FILIAL = SG1.G1_FILIAL
                AND G1_REVFIM IN (SB1.B1_REVATU,'ZZZ','YYY')
                AND SB1.D_E_L_E_T_ = ''
            INNER JOIN SB1010 AS SB12 WITH (NOLOCK) ON G1_COMP = SB12.B1_COD
                AND G1_FILIAL = SB12.B1_FILIAL
                AND SB12.D_E_L_E_T_ = ''
            LEFT JOIN SZ4010 AS SZ4 WITH (NOLOCK) ON SZ4.D_E_L_E_T_ = ''
                AND Z4_FILIAL = SB1.B1_FILIAL
                AND Z4_FAMILIA IN (' ', SB1.B1_XFAMILI)
                AND Z4_GRUPO IN ('', SB1.B1_GRUPO)
            WHERE SG1.D_E_L_E_T_ = ''
                AND G1_FILIAL = '0201'
                AND SB1.B1_XFAMILI IN (Z4_FAMILIA,'') 
                AND SB1.B1_XPORTIF IN (Z4_PORTIFO,'')
                AND SB1.B1_XSBPORT IN (Z4_SBPORTI,'')
                AND SB12.B1_COD = ?
            ORDER BY SB1.B1_VIGENC DESC
    """
    
    # Executar a consulta com o parâmetro
    cursor.execute(query, (sb12_b1_cod,))
    
    if(cursor.rowcount < 0 or cursor.rowcount > 0):
        for row in cursor.fetchall():
            resultados.append(row)
            
        NovaConsulta(resultados)
    else:
        # Crie um objeto Workbook
        wb = openpyxl.Workbook()
        # Crie uma nova planilha
        ws = wb.active
        # Conteúdo antes dos headers
        conteudo_extra = ['PRODUTO PESQUISADO: ' + sb12_b1_cod]
        # Adicione o conteúdo extra na primeira linha
        for i, valor in enumerate(conteudo_extra, start=1):
            ws.cell(row=i, column=1).value = valor
        # Headers
        headers = ['PRODUTO NÃO CONSTA EM ESTRUTURAS']
        # Adicione os headers
        for col, header in enumerate(headers, start=1):
            ws.cell(row=len(conteudo_extra) + 1, column=col).value = header
        
        # Especificar o caminho onde salvar o arquivo
        wb.save(file_path)
        print("Resultados "+nomeArquivo+" exportados para Excel com sucesso!")

def criar_arquivo_fim(diretorio,Caminho_Abrir):
    # Verificar se o diretório existe, caso contrário, criar o diretório
    if not os.path.exists(diretorio):
        os.makedirs(diretorio)
    
    # Especificar o caminho completo do arquivo
    nome_arquivo = os.path.join(diretorio, "FIM.txt")
    
    # Abrir o arquivo em modo de escrita
    with open(nome_arquivo, "w") as arquivo:
        # Escrever conteúdo no arquivo, se necessário
        arquivo.write("Este é o conteúdo do arquivo FIM.txt.\n")
    
    print(f"Arquivo '{nome_arquivo}' criado com sucesso.")
    
    #caminho_absoluto = os.path.abspath(diretorio)
    os.startfile(Caminho_Abrir)

# Especificar o diretório onde o arquivo será salvo
teste = "C:/temp"  # Substitua pelo caminho desejado
file_path = file_path2
# Chamar a função para criar o arquivo no diretório especificado
criar_arquivo_fim(teste,file_path)

conn.close()